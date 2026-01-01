"""
Extract longer-run federal funds rate percentiles from XLSX data files.

This module handles:
- Robustly finding the longer-run FF concept across different sheet structures
- Extracting percentile values (25th, 50th/median, 75th)
- Handling panel splits (SPD/SMP/Combined or Dealer/Participant/Combined)
- Normalizing values to percent format
"""

import re
from pathlib import Path
from typing import List, Dict, Optional, Any, Tuple
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

from .utils import (
    logger,
    XLSX_VALUE_TAG,
    CONCEPT_FF_LONGER_RUN,
    PANEL_SPD,
    PANEL_SMP,
    PANEL_DEALER,
    PANEL_PARTICIPANT,
    PANEL_COMBINED,
    SOURCE_XLSX,
    ExtractedPercentile,
    matches_longer_run_ff,
    normalize_percent,
)


# Value tags to search for (in order of preference)
LONGER_RUN_VALUE_TAGS = [
    "fftr_modalpe_longerrun",
    "fftr_longerrun",
    "fed_funds_longerrun",
    "federal_funds_longerrun",
]

# Aggregation types mapping
AGGREGATION_MAP = {
    "pctl25": "pctl25",
    "pctl_25": "pctl25",
    "p25": "pctl25",
    "25th": "pctl25",
    "pctl50": "pctl50",
    "pctl_50": "pctl50",
    "p50": "pctl50",
    "median": "pctl50",
    "50th": "pctl50",
    "pctl75": "pctl75",
    "pctl_75": "pctl75",
    "p75": "pctl75",
    "75th": "pctl75",
}

# Panel type mapping - normalize all to SPD/SMP/Combined
PANEL_MAP = {
    "combined": PANEL_COMBINED,
    "all": PANEL_COMBINED,
    "total": PANEL_COMBINED,
    "dealer": PANEL_SPD,  # Normalize "Dealer" to "SPD"
    "spd": PANEL_SPD,
    "primary_dealer": PANEL_SPD,
    "participant": PANEL_SMP,  # Normalize "Participant" to "SMP"
    "smp": PANEL_SMP,
    "market_participant": PANEL_SMP,
}


def get_sheet_names(filepath: Path) -> List[str]:
    """Get all sheet names from an Excel file."""
    try:
        wb = load_workbook(filepath, read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
        return sheets
    except Exception as e:
        logger.error(f"Failed to read sheet names from {filepath}: {e}")
        return []


def find_column(df: pd.DataFrame, possible_names: List[str]) -> Optional[str]:
    """Find a column by checking multiple possible names."""
    columns_lower = {str(c).lower(): c for c in df.columns}
    for name in possible_names:
        if name.lower() in columns_lower:
            return columns_lower[name.lower()]
    return None


def parse_long_format(
    df: pd.DataFrame,
    filepath: Path,
    file_url: str,
    survey_date: datetime,
) -> List[ExtractedPercentile]:
    """
    Parse XLSX in long format where:
    - value_tag identifies the concept
    - aggregation identifies pctl25/pctl50/pctl75
    - aggregation_value contains the actual value
    - panel_type (optional) identifies dealer/participant/combined
    """
    results = []
    
    # Find required columns
    value_tag_col = find_column(df, ["value_tag", "valuetag", "tag", "concept", "variable"])
    agg_col = find_column(df, ["aggregation", "agg", "statistic", "stat", "aggregation_type"])
    value_col = find_column(df, ["aggregation_value", "value", "result", "estimate", "number"])
    panel_col = find_column(df, ["panel_type", "panel", "respondent_type", "survey_type"])
    
    if not value_tag_col or not agg_col or not value_col:
        logger.debug(f"Missing required columns. Found: value_tag={value_tag_col}, agg={agg_col}, value={value_col}")
        return results
    
    logger.debug(f"Found columns: value_tag={value_tag_col}, agg={agg_col}, value={value_col}, panel={panel_col}")
    
    # Filter to rows with longer-run FF value_tag
    lr_mask = pd.Series([False] * len(df))
    for tag in LONGER_RUN_VALUE_TAGS:
        lr_mask = lr_mask | (df[value_tag_col].astype(str).str.lower() == tag.lower())
    
    # Also try partial match if exact match fails
    if not lr_mask.any():
        lr_mask = df[value_tag_col].astype(str).str.lower().str.contains(
            r"fftr.*longer|longer.*fftr|fed.*fund.*longer",
            regex=True,
            na=False,
        )
    
    if not lr_mask.any():
        logger.debug(f"No longer-run FF rows found by value_tag in {filepath.name}")
        return results
    
    lr_df = df[lr_mask].copy()
    logger.debug(f"Found {len(lr_df)} longer-run FF rows")
    
    # Group by panel type
    if panel_col:
        panels = lr_df[panel_col].dropna().unique()
    else:
        panels = ["Combined"]
    
    for panel in panels:
        if panel_col:
            panel_df = lr_df[lr_df[panel_col] == panel]
        else:
            panel_df = lr_df
        
        # Normalize panel name
        panel_lower = str(panel).lower().strip()
        panel_name = PANEL_MAP.get(panel_lower, panel)
        
        # Extract percentiles
        pctls = {"pctl25": None, "pctl50": None, "pctl75": None}
        
        for _, row in panel_df.iterrows():
            agg_type = str(row[agg_col]).lower().strip()
            value = row[value_col]
            
            # Map aggregation to our percentile names
            pctl_key = AGGREGATION_MAP.get(agg_type)
            
            if pctl_key and pctls[pctl_key] is None:
                normalized = normalize_percent(value)
                if normalized is not None:
                    pctls[pctl_key] = normalized
        
        # Only add if we found at least one percentile
        if any(v is not None for v in pctls.values()):
            results.append(ExtractedPercentile(
                survey_date=survey_date,
                panel=panel_name,
                concept=CONCEPT_FF_LONGER_RUN,
                pctl25=pctls["pctl25"],
                pctl50=pctls["pctl50"],
                pctl75=pctls["pctl75"],
                source=SOURCE_XLSX,
                file_url=file_url,
                local_path=str(filepath),
            ))
    
    return results


def parse_wide_format(
    df: pd.DataFrame,
    filepath: Path,
    file_url: str,
    survey_date: datetime,
) -> List[ExtractedPercentile]:
    """
    Parse XLSX in wide format where percentiles are in separate columns.
    """
    results = []
    columns = [str(c).lower() for c in df.columns]
    
    # Find value_tag or question_text columns
    value_tag_col = find_column(df, ["value_tag", "valuetag", "tag", "concept"])
    question_text_col = find_column(df, ["question_text", "description", "question", "label"])
    
    # Find rows matching longer-run FF concept
    lr_df = pd.DataFrame()
    
    if value_tag_col:
        for tag in LONGER_RUN_VALUE_TAGS:
            mask = df[value_tag_col].astype(str).str.lower() == tag.lower()
            if mask.any():
                lr_df = df[mask].copy()
                break
    
    if lr_df.empty and question_text_col:
        mask = df[question_text_col].astype(str).apply(matches_longer_run_ff)
        if mask.any():
            lr_df = df[mask].copy()
    
    if lr_df.empty:
        return results
    
    # Find percentile columns
    pctl_cols = {"pctl25": None, "pctl50": None, "pctl75": None}
    
    for col_name in df.columns:
        col_lower = str(col_name).lower()
        if "25" in col_lower or "p25" in col_lower:
            pctl_cols["pctl25"] = col_name
        elif "median" in col_lower or "50" in col_lower or "p50" in col_lower:
            pctl_cols["pctl50"] = col_name
        elif "75" in col_lower or "p75" in col_lower:
            pctl_cols["pctl75"] = col_name
    
    for _, row in lr_df.iterrows():
        pctls = {"pctl25": None, "pctl50": None, "pctl75": None}
        
        for pctl_key, col_name in pctl_cols.items():
            if col_name and col_name in row.index:
                value = row[col_name]
                pctls[pctl_key] = normalize_percent(value)
        
        if any(v is not None for v in pctls.values()):
            results.append(ExtractedPercentile(
                survey_date=survey_date,
                panel=PANEL_COMBINED,
                concept=CONCEPT_FF_LONGER_RUN,
                pctl25=pctls["pctl25"],
                pctl50=pctls["pctl50"],
                pctl75=pctls["pctl75"],
                source=SOURCE_XLSX,
                file_url=file_url,
                local_path=str(filepath),
            ))
    
    return results


def extract_from_xlsx(
    filepath: Path,
    file_url: str,
    survey_date: datetime,
    survey_type: str = "merged",
) -> List[ExtractedPercentile]:
    """
    Extract longer-run federal funds rate percentiles from an XLSX file.
    
    Args:
        filepath: Path to the XLSX file
        file_url: Original URL of the file
        survey_date: Date of the survey
        survey_type: Type of survey (SPD, SMP, or merged)
    
    Returns:
        List of ExtractedPercentile objects
    """
    logger.info(f"Extracting from XLSX: {filepath.name}")
    
    all_results = []
    
    try:
        # Get all sheet names
        sheet_names = get_sheet_names(filepath)
        logger.debug(f"Found sheets: {sheet_names}")
        
        if not sheet_names:
            logger.error(f"No sheets found in {filepath}")
            return [_create_empty_result(filepath, file_url, survey_date, "no_sheets_found")]
        
        # Try each sheet
        for sheet_name in sheet_names:
            try:
                # Read sheet with pandas
                df = pd.read_excel(
                    filepath,
                    sheet_name=sheet_name,
                    engine="openpyxl",
                )
                
                if df.empty:
                    continue
                
                logger.debug(f"Sheet '{sheet_name}': {len(df)} rows, {len(df.columns)} cols")
                
                # Try long format first (most common for NY Fed data)
                results = parse_long_format(df, filepath, file_url, survey_date)
                
                if not results:
                    # Try wide format
                    results = parse_wide_format(df, filepath, file_url, survey_date)
                
                all_results.extend(results)
                
            except Exception as e:
                logger.warning(f"Error parsing sheet '{sheet_name}': {e}")
                continue
        
        # If survey_type indicates SPD or SMP, update panel labels for Combined results
        if survey_type in [PANEL_SPD, PANEL_SMP]:
            for result in all_results:
                if result.panel == PANEL_COMBINED:
                    result.panel = survey_type
        
        # Remove duplicates (same panel)
        seen = set()
        unique_results = []
        for result in all_results:
            key = (result.survey_date, result.panel)
            if key not in seen:
                seen.add(key)
                unique_results.append(result)
        
        if not unique_results:
            logger.warning(f"No percentiles extracted from {filepath.name}")
            return [_create_empty_result(filepath, file_url, survey_date, "question_not_present")]
        
        logger.info(f"Extracted {len(unique_results)} percentile records from {filepath.name}")
        return unique_results
        
    except Exception as e:
        logger.error(f"Failed to parse XLSX {filepath}: {e}")
        return [_create_empty_result(filepath, file_url, survey_date, f"parse_error: {str(e)[:100]}")]


def _create_empty_result(
    filepath: Path,
    file_url: str,
    survey_date: datetime,
    notes: str,
) -> ExtractedPercentile:
    """Create an empty result with notes."""
    return ExtractedPercentile(
        survey_date=survey_date,
        panel=PANEL_COMBINED,
        concept=CONCEPT_FF_LONGER_RUN,
        pctl25=None,
        pctl50=None,
        pctl75=None,
        source=SOURCE_XLSX,
        file_url=file_url,
        local_path=str(filepath),
        notes=notes,
    )


# Export for tests
def find_column_by_patterns(
    columns: List[str],
    patterns: List[str],
    exact_match: bool = False,
) -> Optional[str]:
    """Find a column name matching any of the given patterns."""
    for col in columns:
        col_lower = str(col).lower().strip()
        for pattern in patterns:
            if exact_match:
                if col_lower == pattern.lower():
                    return col
            else:
                if re.search(pattern, col_lower, re.IGNORECASE):
                    return col
    return None


def detect_percentile_columns(
    columns: List[str],
    panel_suffix: Optional[str] = None,
) -> Dict[str, Optional[str]]:
    """
    Detect columns containing percentile values.
    """
    result = {"pctl25": None, "pctl50": None, "pctl75": None}
    
    patterns = {
        "pctl25": [r"p?ctl_?25", r"p25", r"25th"],
        "pctl50": [r"median", r"p?ctl_?50", r"p50", r"50th"],
        "pctl75": [r"p?ctl_?75", r"p75", r"75th"],
    }
    
    for pctl_name, pctl_patterns in patterns.items():
        for col in columns:
            col_lower = str(col).lower()
            for pattern in pctl_patterns:
                if re.search(pattern, col_lower, re.IGNORECASE):
                    if panel_suffix:
                        if panel_suffix.lower() in col_lower:
                            result[pctl_name] = col
                            break
                    else:
                        result[pctl_name] = col
                        break
            if result[pctl_name]:
                break
    
    return result


def extract_percentiles_from_row(
    row: pd.Series,
    columns: Dict[str, str],
) -> Dict[str, Optional[float]]:
    """Extract percentile values from a row."""
    result = {"pctl25": None, "pctl50": None, "pctl75": None}
    
    for pctl_name, col_name in columns.items():
        if col_name and col_name in row.index:
            value = row[col_name]
            result[pctl_name] = normalize_percent(value)
    
    return result


def find_longer_run_rows(
    df: pd.DataFrame,
    value_tag_col: Optional[str] = None,
    question_text_col: Optional[str] = None,
) -> pd.DataFrame:
    """Find rows containing the longer-run federal funds rate concept."""
    if df.empty:
        return pd.DataFrame()
    
    # First try: exact value_tag match
    if value_tag_col and value_tag_col in df.columns:
        for tag in LONGER_RUN_VALUE_TAGS:
            mask = df[value_tag_col].astype(str).str.lower() == tag.lower()
            if mask.any():
                logger.debug(f"Found {mask.sum()} rows via value_tag exact match")
                return df[mask].copy()
        
        # Try partial match
        mask = df[value_tag_col].astype(str).str.lower().str.contains(
            XLSX_VALUE_TAG.lower(),
            na=False,
            regex=False,
        )
        if mask.any():
            logger.debug(f"Found {mask.sum()} rows via value_tag partial match")
            return df[mask].copy()
    
    # Second try: question text keyword matching
    if question_text_col and question_text_col in df.columns:
        mask = df[question_text_col].astype(str).apply(matches_longer_run_ff)
        if mask.any():
            logger.debug(f"Found {mask.sum()} rows via question text match")
            return df[mask].copy()
    
    return pd.DataFrame()


# Keep backward compatible exports
VALUE_TAG_PATTERNS = [
    r"value_?tag",
    r"series_?name",
    r"variable",
    r"concept",
]


if __name__ == "__main__":
    # Test with a sample file
    import sys
    
    if len(sys.argv) > 1:
        test_file = Path(sys.argv[1])
        if test_file.exists():
            results = extract_from_xlsx(
                test_file,
                file_url="test://example.xlsx",
                survey_date=datetime(2024, 12, 1),
            )
            
            for r in results:
                print(f"\nPanel: {r.panel}")
                print(f"  25th: {r.pctl25}")
                print(f"  Median: {r.pctl50}")
                print(f"  75th: {r.pctl75}")
                if r.notes:
                    print(f"  Notes: {r.notes}")
